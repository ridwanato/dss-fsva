import os
import sys
import json
import zipfile
import openpyxl

def generate_export(level, json_data_path, output_dir, form_requested):
    with open(json_data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # data contains list of rows (raw_indicators + fsva_results + geometry)
    rows_data = data.get('rows', [])
    tahun = data.get('tahun', '2026')
    kabupaten = data.get('kabupaten', 'DAERAH')

    base_template_dir = r"c:\Users\THINKPAD\.gemini\antigravity\scratch\dss-fsva\public\templates_v2"
    template_folder = os.path.join(base_template_dir, 'provinsi' if level == 'provinsi' else 'kab_kota')

    os.makedirs(output_dir, exist_ok=True)

    generated_files = {}

    # File mapping
    if level == 'provinsi':
        file_map = {
            '0': "0. Form Validasi Data FSVA Prov 2026 ver.2.xlsx",
            '1': "1. Form Hitung NCPR FSVA Prov 2026 ver.2.xlsx",
            '2': "2. Form Penentuan Cut Off dan Analisis Komposit Baseline FSVA Prov ver.2.xlsx",
            '3': "3. Form Layout Hasil Baru Menyusun FSVA Prov 2026 Vers.2.xlsx",
            'template': "Template Hasil UPDATE FSVA Prov 2026 ke Peta.xlsx"
        }
    else:
        file_map = {
            '0': "0. Form Validasi Data FSVA Kabupaten Kota 2026 ver.2.xlsx",
            '1': "1. Form Hitung NCPR FSVA Kabupaten Kota 2026 ver.2.xlsx",
            '2': "2. Form Penentuan Cut Off dan Analisis Komposit Baseline FSVA Kabupaten Kota ver.2.xlsx",
            '3': "3. Form Layout Hasil Baru Menyusun FSVA Kabupaten Kota 2026 Vers.2.xlsx",
            'template': "Template Hasil UPDATE FSVA KabKota 2026 ke Peta.xlsx"
        }

    forms_to_process = list(file_map.keys()) if form_requested == 'zip' or form_requested == 'all' else [form_requested]

    for f_key in forms_to_process:
        if f_key not in file_map:
            continue
        template_name = file_map[f_key]
        src_path = os.path.join(template_folder, template_name)
        out_path = os.path.join(output_dir, template_name)

        if not os.path.exists(src_path):
            continue

        wb = openpyxl.load_workbook(src_path, data_only=False)

        # ----------------------------------------------------
        # FORM 0 POPULATION
        # ----------------------------------------------------
        if f_key == '0':
            # Sheet 0.1 Produksi Pangan & Penduduk
            if '0.1 Produksi Pangan & Penduduk' in wb.sheetnames:
                ws = wb['0.1 Produksi Pangan & Penduduk']
                for idx, r in enumerate(rows_data, start=6):
                    ws.cell(idx, 1, idx - 5)
                    ws.cell(idx, 2, r.get('nama_kecamatan', r.get('nama_kabupaten', '')))
                    ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', r.get('kode_bps', '')))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('nama_desa', r.get('nama_kabupaten', '')))
                    
                    ws.cell(idx, 10, r.get('produksi_padi', 0))
                    ws.cell(idx, 19, r.get('produksi_jagung', 0))
                    ws.cell(idx, 25, r.get('produksi_ubi_kayu', 0))
                    ws.cell(idx, 31, r.get('produksi_ubi_jalar', 0))
                    ws.cell(idx, 37, r.get('produksi_sagu', 0))
                    ws.cell(idx, 43, r.get('produksi_pisang', 0))
                    ws.cell(idx, 49, r.get('jumlah_penduduk', 0))

            # Sheet 0.2 Konsumsi Energi
            if '0.2 Konsumsi Energi' in wb.sheetnames:
                ws = wb['0.2 Konsumsi Energi']
                for idx, r in enumerate(rows_data, start=6):
                    ws.cell(idx, 1, idx - 5)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('nama_desa', ''))
                    ws.cell(idx, 10, r.get('konsumsi_energi', 0))

            # Sheet 0.3 Konsumsi Protein Hewani
            if '0.3 Konsumsi Protein Hewani' in wb.sheetnames:
                ws = wb['0.3 Konsumsi Protein Hewani']
                for idx, r in enumerate(rows_data, start=6):
                    ws.cell(idx, 1, idx - 5)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('nama_desa', ''))
                    ws.cell(idx, 10, r.get('konsumsi_protein', 0))

            # Sheet 0.4 Stok_Cadangan
            if '0.4 Stok_Cadangan' in wb.sheetnames:
                ws = wb['0.4 Stok_Cadangan']
                for idx, r in enumerate(rows_data, start=6):
                    ws.cell(idx, 1, idx - 5)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('nama_desa', ''))
                    ws.cell(idx, 10, r.get('cbpk', 0))
                    ws.cell(idx, 19, r.get('jumlah_penduduk_kab', r.get('jumlah_penduduk', 0)))
                    ws.cell(idx, 25, r.get('cadangan_cbpd', 0))
                    ws.cell(idx, 30, r.get('cadangan_lpm', 0))

            # Sheet 0.5 Indikator Keterjangkauan
            if '0.5 Indikator Keterjangkauan' in wb.sheetnames:
                ws = wb['0.5 Indikator Keterjangkauan']
                for idx, r in enumerate(rows_data, start=6):
                    ws.cell(idx, 1, idx - 5)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('nama_desa', ''))
                    ws.cell(idx, 10, r.get('pct_miskin', 0))
                    ws.cell(idx, 16, r.get('cv_harga_beras', 0))
                    ws.cell(idx, 19, r.get('cv_harga_ayam', 0))
                    ws.cell(idx, 22, r.get('cv_harga_telur', 0))
                    ws.cell(idx, 25, r.get('cv_harga_minyak', 0))
                    ws.cell(idx, 36, r.get('pou', 0))

            # Sheet 0.6 Indikator Pemanfaatan
            if '0.6 Indikator Pemanfaatan' in wb.sheetnames:
                ws = wb['0.6 Indikator Pemanfaatan']
                for idx, r in enumerate(rows_data, start=6):
                    ws.cell(idx, 1, idx - 5)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('nama_desa', ''))
                    ws.cell(idx, 10, r.get('lama_sekolah_perempuan', 0))
                    ws.cell(idx, 16, r.get('pct_no_water', 0))
                    ws.cell(idx, 19, r.get('skor_pph', 0))
                    ws.cell(idx, 25, r.get('pct_stunting', 0))

        # ----------------------------------------------------
        # FORM 1 POPULATION
        # ----------------------------------------------------
        elif f_key == '1':
            padi_sheets = [s for s in wb.sheetnames if 'Padi' in s]
            if padi_sheets:
                ws = wb[padi_sheets[0]]
                for idx, r in enumerate(rows_data, start=5):
                    ws.cell(idx, 1, idx - 4)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('nama_desa', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('produksi_padi', 0))

            jagung_sheets = [s for s in wb.sheetnames if 'Jagung' in s]
            if jagung_sheets:
                ws = wb[jagung_sheets[0]]
                for idx, r in enumerate(rows_data, start=5):
                    ws.cell(idx, 1, idx - 4)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('nama_desa', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('produksi_jagung', 0))

            ubi_k_sheets = [s for s in wb.sheetnames if 'Ubi Kayu' in s]
            if ubi_k_sheets:
                ws = wb[ubi_k_sheets[0]]
                for idx, r in enumerate(rows_data, start=5):
                    ws.cell(idx, 1, idx - 4)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('nama_desa', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('produksi_ubi_kayu', 0))

            ubi_j_sheets = [s for s in wb.sheetnames if 'Ubi Jalar' in s]
            if ubi_j_sheets:
                ws = wb[ubi_j_sheets[0]]
                for idx, r in enumerate(rows_data, start=5):
                    ws.cell(idx, 1, idx - 4)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('nama_desa', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('produksi_ubi_jalar', 0))

            sagu_sheets = [s for s in wb.sheetnames if 'Sagu' in s]
            if sagu_sheets:
                ws = wb[sagu_sheets[0]]
                for idx, r in enumerate(rows_data, start=5):
                    ws.cell(idx, 1, idx - 4)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('nama_desa', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('produksi_sagu', 0))

            pisang_sheets = [s for s in wb.sheetnames if 'Pisang' in s]
            if pisang_sheets:
                ws = wb[pisang_sheets[0]]
                for idx, r in enumerate(rows_data, start=5):
                    ws.cell(idx, 1, idx - 4)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('nama_desa', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('produksi_pisang', 0))

            ncpr_sheets = [s for s in wb.sheetnames if 'NCPR' in s]
            if ncpr_sheets:
                ws = wb[ncpr_sheets[0]]
                for idx, r in enumerate(rows_data, start=3):
                    ws.cell(idx, 1, idx - 2)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('nama_desa', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))

        # ----------------------------------------------------
        # FORM 2 POPULATION
        # ----------------------------------------------------
        elif f_key == '2':
            if '2.1 Data FSVA 2025 & Bobot' in wb.sheetnames:
                ws = wb['2.1 Data FSVA 2025 & Bobot']
                for idx, r in enumerate(rows_data, start=7):
                    ws.cell(idx, 1, idx - 6)
                    ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                    ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                    ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                    ws.cell(idx, 5, r.get('kode_bps', ''))
                    ws.cell(idx, 6, r.get('nama_desa', ''))
                    ws.cell(idx, 7, r.get('ncpr', 0))
                    ws.cell(idx, 9, r.get('pct_ake', 0))
                    ws.cell(idx, 11, r.get('pct_prohe', 0))
                    ws.cell(idx, 13, r.get('rasio_cadangan', 0))
                    ws.cell(idx, 15, r.get('pct_miskin', 0))
                    ws.cell(idx, 17, r.get('cv_harga', 0))
                    ws.cell(idx, 19, r.get('pou', 0))
                    ws.cell(idx, 21, r.get('lama_sekolah', 0))
                    ws.cell(idx, 23, r.get('pct_no_water', 0))
                    ws.cell(idx, 25, r.get('skor_pph', 0))
                    ws.cell(idx, 27, r.get('pct_stunting', 0))

        # ----------------------------------------------------
        # FORM 3 POPULATION
        # ----------------------------------------------------
        elif f_key == '3':
            sheet3_name = '3.2 Data & Hasil FSVA 2026' if '3.2 Data & Hasil FSVA 2026' in wb.sheetnames else wb.sheetnames[1]
            ws = wb[sheet3_name]
            for idx, r in enumerate(rows_data, start=7):
                ws.cell(idx, 1, idx - 6)
                ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                ws.cell(idx, 5, r.get('kode_bps', ''))
                ws.cell(idx, 6, r.get('nama_desa', ''))
                ws.cell(idx, 7, r.get('produksi_padi', 0))
                ws.cell(idx, 8, r.get('produksi_jagung', 0))
                ws.cell(idx, 9, r.get('produksi_ubi_kayu', 0))
                ws.cell(idx, 10, r.get('produksi_ubi_jalar', 0))
                ws.cell(idx, 11, r.get('produksi_sagu', 0))
                ws.cell(idx, 12, r.get('produksi_pisang', 0))
                ws.cell(idx, 13, r.get('jumlah_penduduk', 0))
                ws.cell(idx, 14, r.get('konsumsi_energi', 0))
                ws.cell(idx, 15, r.get('konsumsi_protein', 0))
                ws.cell(idx, 16, r.get('cadangan_cbpd', 0) + r.get('cadangan_lpm', 0))
                ws.cell(idx, 17, r.get('pct_miskin', 0))
                ws.cell(idx, 18, r.get('cv_harga_beras', 0))
                ws.cell(idx, 19, r.get('cv_harga_ayam', 0))
                ws.cell(idx, 20, r.get('cv_harga_telur', 0))
                ws.cell(idx, 21, r.get('cv_harga_minyak', 0))
                ws.cell(idx, 22, r.get('pou', 0))
                ws.cell(idx, 23, r.get('lama_sekolah_perempuan', 0))
                ws.cell(idx, 24, r.get('pct_no_water', 0))
                ws.cell(idx, 25, r.get('skor_pph', 0))
                ws.cell(idx, 26, r.get('pct_stunting', 0))

        # ----------------------------------------------------
        # TEMPLATE PETA POPULATION
        # ----------------------------------------------------
        elif f_key == 'template':
            ws = wb.worksheets[0]
            for idx, r in enumerate(rows_data, start=2):
                ws.cell(idx, 1, idx - 1)
                ws.cell(idx, 2, r.get('nama_kecamatan', ''))
                ws.cell(idx, 3, r.get('kode_kecamatan', ''))
                ws.cell(idx, 4, r.get('kode_kemendagri', ''))
                ws.cell(idx, 5, r.get('kode_bps', ''))
                ws.cell(idx, 6, r.get('nama_desa', ''))
                ws.cell(idx, 7, r.get('p_ncpr', 6))
                ws.cell(idx, 8, r.get('p_energy', 6))
                ws.cell(idx, 9, r.get('p_protein', 6))
                ws.cell(idx, 10, r.get('p_cadangan', 6))
                ws.cell(idx, 11, r.get('p_poverty', 6))
                ws.cell(idx, 12, r.get('p_cv_harga', 6))
                ws.cell(idx, 13, r.get('p_pou', 6))
                ws.cell(idx, 14, r.get('p_sekolah', 6))
                ws.cell(idx, 15, r.get('p_air', 6))
                ws.cell(idx, 16, r.get('p_pph', 6))
                ws.cell(idx, 17, r.get('p_stunting', 4))
                ws.cell(idx, 18, r.get('indeks_komposit', 0))
                ws.cell(idx, 20, r.get('prioritas', 6))

        wb.save(out_path)
        generated_files[f_key] = out_path

    # If ZIP requested
    if form_requested == 'zip':
        zip_filename = f"Paket_Form_FSVA_V2_{level}_{kabupaten}_{tahun}.zip".replace(" ", "_")
        zip_path = os.path.join(output_dir, zip_filename)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f_key, filepath in generated_files.items():
                arcname = os.path.basename(filepath)
                zf.write(filepath, arcname)
        print(json.dumps({"success": True, "file_path": zip_path, "zip_name": zip_filename}))
    else:
        file_path = generated_files.get(form_requested, '')
        print(json.dumps({"success": True, "file_path": file_path, "file_name": os.path.basename(file_path)}))

if __name__ == '__main__':
    if len(sys.argv) < 5:
        print(json.dumps({"success": False, "error": "Invalid arguments"}))
        sys.exit(1)
    
    level = sys.argv[1]
    json_path = sys.argv[2]
    out_dir = sys.argv[3]
    form_req = sys.argv[4]

    try:
        generate_export(level, json_path, out_dir, form_req)
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
